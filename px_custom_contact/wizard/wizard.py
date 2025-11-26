# import base64
# import io
# import csv
# import xlrd
# import openpyxl
# import os
# from odoo import models, fields, _
# from odoo.exceptions import ValidationError
#
# class ImportExcelWizard(models.TransientModel):
#     _name = 'import.excel.wizard'
#     _description = 'Import Excel Wizard'
#
#     file = fields.Binary("Excel File", required=True)
#     filename = fields.Char("Filename")
#
#
#
#     def import_contacts_from_excel(self):
#         if not self.file:
#             raise ValidationError(_("Please upload an Excel file."))
#
#         # data = base64.b64decode(self.file)
#         # file_stream = io.BytesIO(data)
#         # workbook = openpyxl.load_workbook(file_stream)
#         # sheet = workbook.active
#         #
#         # raw_headers = [str(cell.value).strip() if cell.value else '' for cell in sheet[1]]
#         # print(raw_headers,"raw headers")
#         data = base64.b64decode(self.file)
#
#                 # ---------------- Detect Extension Safely ----------------
#         file_ext = (os.path.splitext(self.filename or '')[1] or '').lower().strip()
#
#         # 🟢 DEBUG INFO
#         print("DEBUG >>> filename:", self.filename)
#         print("DEBUG >>> file_ext:", file_ext)
#
#         # ---------------- CSV ----------------
#         if file_ext == ".csv":
#             try:
#                 file_stream = io.StringIO(data.decode("utf-8"))
#                 reader = csv.reader(file_stream)
#                 rows = list(reader)
#                 if not rows:
#                     raise ValidationError(_("CSV file is empty."))
#                 raw_headers = [str(h).strip() for h in rows[0]]
#                 sheet = rows[1:]
#             except Exception as e:
#                 raise ValidationError(_("Invalid CSV file: %s") % str(e))
#
#         # ---------------- XLS ----------------
#         elif file_ext == ".xls":
#             try:
#                 book = xlrd.open_workbook(file_contents=data)
#                 sheet_xls = book.sheet_by_index(0)
#                 raw_headers = [str(sheet_xls.cell_value(0, col)).strip()
#                                for col in range(sheet_xls.ncols)]
#                 sheet = []
#                 for row_idx in range(1, sheet_xls.nrows):
#                     sheet.append([sheet_xls.cell_value(row_idx, col)
#                                   for col in range(sheet_xls.ncols)])
#             except Exception as e:
#                 raise ValidationError(_("Invalid XLS file: %s") % str(e))
#
#         # ---------------- XLSX ----------------
#         elif file_ext == ".xlsx":
#             try:
#                 file_stream = io.BytesIO(data)
#                 workbook = openpyxl.load_workbook(file_stream)
#                 sheet_obj = workbook.active
#                 raw_headers = [str(cell.value).strip() if cell.value else '' for cell in sheet_obj[1]]
#                 sheet = []
#                 for row in sheet_obj.iter_rows(min_row=2, values_only=True):
#                     sheet.append(list(row))  # هنا sheet بقت list of lists
#             except Exception as e:
#                 raise ValidationError(_("Invalid XLSX file: %s") % str(e))
#         # ---------------- Unsupported ----------------
#         else:
#             raise ValidationError(_("Unsupported file format. Please upload a CSV, XLS, or XLSX file."))
#
#
#
#         SMART_FIELD_MAPPING = {
#             'name': ['name',  'display name', 'opportunity'],
#             'email_from': ['email', 'e-mail', 'email address'],
#             'phone': ['phone', 'phone number'],
#             'mobile': ['mobile', 'mobile number'],
#             'expected_revenue': ['expected revenue', 'revenue', 'amount'],
#             'user_id': ['salesperson', 'sales person', 'sales rep', 'assigned to'],
#             'stage_id': ['stage', 'status'],
#             'partner_id': ['company', 'customer', 'client'],
#             'contact_name': ['contact', 'contact name'],
#         }
#
#         def map_header_to_field(header, existing_fields):
#             header_clean = header.strip().lower()
#             for field, keywords in SMART_FIELD_MAPPING.items():
#                 if header_clean in [kw.lower() for kw in keywords]:
#                     if field not in existing_fields:
#                         return field
#                     else:
#                         return header_clean
#             return header_clean
#
#         headers = []
#         for h in raw_headers:
#             field = map_header_to_field(h, headers)
#             headers.append(field)
#
#
#         Lead = self.env['crm.lead']
#         valid_fields = Lead.fields_get().keys()
#
#         inserted = 0
#         skipped = 0
#         print(valid_fields,"valid fields")
#         for i, row in enumerate(sheet, start=2):  # sheet هنا عبارة عن list of rows في كل الحالات
#             values = {}
#             for col_idx, cell in enumerate(row):
#                 if col_idx < len(headers):
#                     field_name = headers[col_idx]
#                     if field_name in valid_fields:
#                         if cell is not None:
#                             values[field_name] = str(cell).strip()
#                         else:
#                             values[field_name] = ''
#
#             print(values.get('name'),"vals before")
#
#             if not values.get('name') and (values.get('contact_name') or values.get('partner_id')):
#                 values['name'] = values.get('contact_name') or values.get('partner_id')
#             if not values.get('name'):
#                 continue
#             print(values.get('name'),"vals")
#             existing = Lead.search([
#                 ('name', '=', values.get('name')),
#                 ('phone', '=', values.get('phone')),
#                 ('mobile', '=', values.get('mobile')),
#                 ('type', '=', 'opportunity')
#             ], limit=1)
#             print(">>>>>>>>>>>>>",existing)
#             if existing:
#                 existing.write({'duplicate': True})
#                 existing.write({'color': 1})
#                 skipped += 1
#                 continue
#
#             if values.get('user_id'):
#                 print(values.get('user_id'))
#                 user = self.env['res.users'].search([('name', '=', values['user_id'])], limit=1)
#                 if user:
#                     values['user_id'] = user.id
#                 else:
#                     values['user_id'] = False
#
#                     # values.pop('user_id')
#
#             if values.get('stage_id'):
#                 stage = self.env['crm.stage'].search([('name', '=', values['stage_id'])], limit=1)
#                 if stage:
#                     values['stage_id'] = stage.id
#                 else:
#                     values['stage_id'] = False
#
#                     # values.pop('stage_id')
#
#
#             if values.get('partner_id'):
#                 print(values.get('partner_id'))
#                 partner = self.env['res.partner'].search([('name', '=', values['partner_id'])], limit=1)
#                 if partner:
#                     print(partner,"from inside if")
#                     values['partner_id'] = partner.id
#                 else:
#                     print("else")
#                     values['partner_id']=False
#             else:
#                 if values.get('contact_name'):
#                     partner = self.env['res.partner'].search([('name', '=', values['contact_name'])], limit=1)
#                     if partner:
#                         print(partner, "from inside if 1")
#                         values['partner_id'] = partner.id
#                     else:
#                         values['partner_id'] = False
#                         print("else 1")
#                         # values.pop('partner_id')
#                 if values.get('name'):
#                     partner = self.env['res.partner'].search([('name', '=', values['name'])], limit=1)
#                     if partner:
#                         print(partner, "from inside if 1")
#                         values['partner_id'] = partner.id
#                     else:
#                         values['partner_id'] = False
#                         print("else 1")
#                         # values.pop('partner_id')
#
#             # print(values['partner_id'])
#             values['type'] = 'opportunity'
#             values['duplicate'] = False
#             print(">>>>>>>>>>>>>",values)
#             Lead.create(values)
#             inserted += 1
#
#         return {
#             'type': 'ir.actions.client',
#             'tag': 'display_notification',
#             'params': {
#                 'title': _('Import Done'),
#                 'message': _("Added: %s | Duplicates: %s" % (inserted, skipped)),
#                 'type': 'success',
#                 'sticky': True,
#             }
#         }
#
#


import base64
import io
import csv
import xlrd
import openpyxl
import os
from odoo import models, fields, _
from odoo.exceptions import ValidationError


class ImportExcelWizard(models.TransientModel):
    _name = 'import.excel.wizard'
    _description = 'Import Excel Wizard'

    file = fields.Binary("Excel File", required=True)
    filename = fields.Char("Filename")

    def import_contacts_from_excel(self):
        if not self.file:
            raise ValidationError(_("Please upload an Excel file."))

        data = base64.b64decode(self.file)
        file_ext = (os.path.splitext(self.filename or '')[1] or '').lower().strip()

        # ---------------- CSV ----------------
        if file_ext == ".csv":
            try:
                file_stream = io.StringIO(data.decode("utf-8"))
                reader = csv.reader(file_stream)
                rows = list(reader)
                if not rows:
                    raise ValidationError(_("CSV file is empty."))
                raw_headers = [str(h).strip() for h in rows[0]]
                sheet = rows[1:]
            except Exception as e:
                raise ValidationError(_("Invalid CSV file: %s") % str(e))

        # ---------------- XLS ----------------
        elif file_ext == ".xls":
            try:
                book = xlrd.open_workbook(file_contents=data)
                sheet_xls = book.sheet_by_index(0)
                raw_headers = [str(sheet_xls.cell_value(0, col)).strip()
                               for col in range(sheet_xls.ncols)]
                sheet = []
                for row_idx in range(1, sheet_xls.nrows):
                    sheet.append([sheet_xls.cell_value(row_idx, col)
                                  for col in range(sheet_xls.ncols)])
            except Exception as e:
                raise ValidationError(_("Invalid XLS file: %s") % str(e))

        # ---------------- XLSX ----------------
        elif file_ext == ".xlsx":
            try:
                file_stream = io.BytesIO(data)
                workbook = openpyxl.load_workbook(file_stream)
                sheet_obj = workbook.active
                raw_headers = [str(cell.value).strip() if cell.value else '' for cell in sheet_obj[1]]
                sheet = []
                for row in sheet_obj.iter_rows(min_row=2, values_only=True):
                    sheet.append(list(row))
            except Exception as e:
                raise ValidationError(_("Invalid XLSX file: %s") % str(e))

        # ---------------- Unsupported ----------------
        else:
            raise ValidationError(_("Unsupported file format. Please upload a CSV, XLS, or XLSX file."))

        # ---------------- Field Mapping ----------------
        SMART_FIELD_MAPPING = {
            'name': ['name', 'display name', 'opportunity'],
            'email_from': ['email', 'e-mail', 'email address'],
            'phone': ['phone', 'phone number'],
            'mobile': ['mobile', 'mobile number'],
            'expected_revenue': ['expected revenue', 'revenue', 'amount'],
            'user_id': ['salesperson', 'sales person', 'sales rep', 'assigned to'],
            'stage_id': ['stage', 'status'],
            'partner_id': ['company', 'customer', 'client'],
            'contact_name': ['contact', 'contact name'],
        }

        def map_header_to_field(header, existing_fields):
            header_clean = header.strip().lower()
            for field, keywords in SMART_FIELD_MAPPING.items():
                if header_clean in [kw.lower() for kw in keywords]:
                    if field not in existing_fields:
                        return field
                    else:
                        return header_clean
            return header_clean

        headers = [map_header_to_field(h, []) for h in raw_headers]

        Lead = self.env['crm.lead']
        valid_fields = Lead.fields_get().keys()

        inserted, skipped, no_name , duplicate= 0, 0, 0, 0
        seen_in_file = set()

        for i, row in enumerate(sheet, start=2):
            values = {}
            for col_idx, cell in enumerate(row):
                if col_idx < len(headers):
                    field_name = headers[col_idx]
                    if field_name in valid_fields:
                        values[field_name] = str(cell).strip() if cell else ''

            # ---------------- "no name" ----------------
            if not values.get('name') and (values.get('contact_name') or values.get('partner_id')):
                values['name'] = values.get('contact_name') or values.get('partner_id')

            if not values.get('name') or not values['name'].strip():
                no_name += 1
                continue

            # ---------------- Duplicate Handling ----------------
            unique_key = (values.get('name').strip().lower(),
                          values.get('phone'),
                          values.get('mobile'))

            if unique_key in seen_in_file:
                skipped += 1
                continue
            else:
                seen_in_file.add(unique_key)

            # --------- Search Duplicate in DB ----------
            domain = []
            if values.get('name'):
                domain.append(('name', '=', values['name'].strip()))
            if values.get('phone'):
                domain.append(('phone', '=', values['phone']))
            if values.get('mobile'):
                domain.append(('mobile', '=', values['mobile']))

            if domain:
                existing = Lead.search(['|'] * (len(domain) - 1) + domain, limit=1)
            else:
                existing = False


            # existing = Lead.search(domain, limit=1)
            if existing:
                existing.write({'duplicate': True, 'color': 1})
                duplicate += 1
                continue

            # Resolve relations
            if values.get('user_id'):
                # try login
                user = self.env['res.users'].search([('login', '=', values['user_id'])], limit=1)
                if not user:
                    # fallback to name
                    user = self.env['res.users'].search([('name', '=', values['user_id'])], limit=1)
                values['user_id'] = user.id if user else False

            if values.get('stage_id'):
                stage = self.env['crm.stage'].search([('name', '=', values['stage_id'])], limit=1)
                values['stage_id'] = stage.id if stage else False

            if values.get('partner_id'):
                partner = self.env['res.partner'].search([('name', '=', values['partner_id'])], limit=1)
                values['partner_id'] = partner.id if partner else False
            elif values.get('contact_name'):
                partner = self.env['res.partner'].search([('name', '=', values['contact_name'])], limit=1)
                values['partner_id'] = partner.id if partner else False
            elif values.get('name'):
                partner = self.env['res.partner'].search([('name', '=', values['name'])], limit=1)
                values['partner_id'] = partner.id if partner else False

            values['type'] = 'opportunity'
            values['duplicate'] = False
            Lead.create(values)
            inserted += 1

        # ---------------- Result Notification ----------------
        message = _("Added: %s | Skipped (file dup): %s | Duplicates in DB: %s") % (
            inserted, skipped, duplicate
        )

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Import Done'),
                'message': message,
                'type': 'success',
                'sticky': True,
            }
        }
